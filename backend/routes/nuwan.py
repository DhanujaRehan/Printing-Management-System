"""
Nuwan Dashboard routes — Branch overview for management monitoring.
Role: 'nuwan' — read-only executive dashboard for branch print monitoring.
"""

from fastapi import APIRouter, HTTPException, Depends, Query
from pydantic import BaseModel
from typing import Optional
from db.database import query
from middleware.auth import get_current_user, require_role
from datetime import date, timedelta

router = APIRouter(prefix="/api/nuwan", tags=["Nuwan Dashboard"])


def require_nuwan(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ("nuwan", "manager", "dba"):
        raise HTTPException(status_code=403, detail="Access denied")
    return current_user



# ── Branches list for Nuwan filters ─────────────────────────────────────────
@router.get("/branches")
def get_branches_for_nuwan(current_user: dict = Depends(require_nuwan)):
    """All active branches — for filter dropdowns."""
    return query("""
        SELECT id, code, name FROM branches
        WHERE is_active = TRUE ORDER BY name
    """) or []

# ── Toner levels per branch ──────────────────────────────────────────────────
# ============================================================
# Replace ONLY the get_toner_levels function in nuwan.py
# Find the function starting with:
#   @router.get("/toner")
#   def get_toner_levels
# and replace the entire function with this:
# ============================================================

@router.get("/toner")
def get_toner_levels(current_user: dict = Depends(require_nuwan)):
    """
    Toner levels — uses total_copies_done from toner_installations directly.
    This is set correctly when toner is installed or manually corrected.
    Daily print_count updates it via the print-log POST endpoint.
    """
    rows = query("""
        SELECT
            b.id                    AS branch_id,
            b.code                  AS branch_code,
            b.name                  AS branch_name,
            p.id                    AS printer_id,
            p.printer_code,
            p.model,
            tm.model_code           AS toner_model_code,
            COALESCE(ti.yield_copies, 0)        AS yield_copies,
            COALESCE(ti.total_copies_done, 0)   AS total_copies_done,
            -- Remaining copies
            GREATEST(0,
                COALESCE(ti.yield_copies, 0) - COALESCE(ti.total_copies_done, 0)
            ) AS copies_remaining,
            -- Percentage remaining  (uses stored total_copies_done — no SUM recalculation)
            ROUND(GREATEST(0.0,
                100.0 * (COALESCE(ti.yield_copies, 0) - COALESCE(ti.total_copies_done, 0))
                / NULLIF(ti.yield_copies, 0)
            ), 1) AS current_pct,
            -- Days remaining based on avg_daily_copies
            CASE
                WHEN COALESCE(ti.avg_daily_copies, 0) > 0 THEN
                    ROUND(
                        GREATEST(0,
                            COALESCE(ti.yield_copies, 0) - COALESCE(ti.total_copies_done, 0)
                        )::NUMERIC / ti.avg_daily_copies
                    )
                ELSE NULL
            END AS days_remaining,
            -- Status
            CASE
                WHEN ti.id IS NULL THEN 'unknown'
                WHEN ROUND(GREATEST(0.0, 100.0*(COALESCE(ti.yield_copies,0)-COALESCE(ti.total_copies_done,0))/NULLIF(ti.yield_copies,0)),1) <= 10  THEN 'critical'
                WHEN ROUND(GREATEST(0.0, 100.0*(COALESCE(ti.yield_copies,0)-COALESCE(ti.total_copies_done,0))/NULLIF(ti.yield_copies,0)),1) <= 25  THEN 'low'
                WHEN ROUND(GREATEST(0.0, 100.0*(COALESCE(ti.yield_copies,0)-COALESCE(ti.total_copies_done,0))/NULLIF(ti.yield_copies,0)),1) <= 50  THEN 'medium'
                ELSE 'good'
            END AS status
        FROM printers p
        JOIN branches b ON b.id = p.branch_id AND b.is_active = TRUE
        LEFT JOIN toner_installations ti ON ti.printer_id = p.id AND ti.is_current = TRUE
        LEFT JOIN toner_models tm ON tm.id = ti.toner_model_id
        WHERE p.is_active = TRUE
        ORDER BY current_pct ASC NULLS LAST, b.code, p.printer_code
    """) or []
    return rows


# ── Yesterday's print totals per branch ─────────────────────────────────────
@router.get("/prints/yesterday")
def get_yesterday_prints(current_user: dict = Depends(require_nuwan)):
    """Total prints logged yesterday, per branch. Also shows which branches did NOT log."""
    yesterday = (date.today() - timedelta(days=1)).isoformat()

    logged = query("""
        SELECT
            b.id            AS branch_id,
            b.code          AS branch_code,
            b.name          AS branch_name,
            COALESCE(SUM(pl.print_count), 0)    AS total_prints,
            COUNT(DISTINCT pl.printer_id)        AS printers_logged,
            COALESCE(SUM(pl.a4_single),0)
              + COALESCE(SUM(pl.a4_double),0)    AS a4_total,
            COALESCE(SUM(pl.b4_single),0)
              + COALESCE(SUM(pl.b4_double),0)    AS b4_total,
            COALESCE(SUM(pl.letter_single),0)
              + COALESCE(SUM(pl.letter_double),0) AS legal_total,
            TRUE                                 AS has_submitted
        FROM branches b
        JOIN printers p ON p.branch_id = b.id AND p.is_active = TRUE
        JOIN print_logs pl ON pl.printer_id = p.id AND pl.log_date = %s::date
        WHERE b.is_active = TRUE
        GROUP BY b.id, b.code, b.name
        ORDER BY b.code
    """, (yesterday,))

    all_branches = query("""
        SELECT
            b.id   AS branch_id,
            b.code AS branch_code,
            b.name AS branch_name
        FROM branches b
        WHERE b.is_active = TRUE
        ORDER BY b.code
    """)

    logged_ids = {r["branch_id"] for r in (logged or [])}
    logged_map = {r["branch_id"]: r for r in (logged or [])}

    result = []
    for b in (all_branches or []):
        if b["branch_id"] in logged_ids:
            result.append(logged_map[b["branch_id"]])
        else:
            result.append({
                "branch_id":       b["branch_id"],
                "branch_code":     b["branch_code"],
                "branch_name":     b["branch_name"],
                "total_prints":    0,
                "printers_logged": 0,
                "a4_total":        0,
                "b4_total":        0,
                "legal_total":     0,
                "has_submitted":   False,
            })

    return {
        "date":          yesterday,
        "grand_total":   sum(r["total_prints"] for r in result),
        "branches":      result,
        "missing_count": sum(1 for r in result if not r["has_submitted"]),
    }

# ════════════════════════════════════════════════════════════
# ADD THIS ENDPOINT TO backend/routes/nuwan.py
# Paste it AFTER the get_yesterday_prints function
# ════════════════════════════════════════════════════════════

@router.get("/prints/paper-summary")
def get_paper_summary(
    branch_id: int = None,
    date_from: str = None,
    date_to:   str = None,
    month:     int = None,
    year:      int = None,
    current_user: dict = Depends(require_nuwan)
):
    """Daily paper totals from daily_paper_logs table — for Nuwan & Manager dashboards."""
    filters = ["1=1"]
    params  = []
    if branch_id:
        filters.append("dpl.branch_id = %s"); params.append(branch_id)
    if date_from:
        filters.append("dpl.log_date >= %s::date"); params.append(date_from)
    if date_to:
        filters.append("dpl.log_date <= %s::date"); params.append(date_to)
    if month and year:
        filters.append("EXTRACT(MONTH FROM dpl.log_date) = %s")
        filters.append("EXTRACT(YEAR  FROM dpl.log_date) = %s")
        params.extend([month, year])
    elif year:
        filters.append("EXTRACT(YEAR FROM dpl.log_date) = %s"); params.append(year)

    where = " AND ".join(filters)

    rows = query(f"""
        SELECT
            dpl.log_date,
            b.id   AS branch_id,
            b.code AS branch_code,
            b.name AS branch_name,
            dpl.paper_type,
            dpl.single_side,
            dpl.double_side,
            dpl.single_side + dpl.double_side AS total_sheets,
            u.full_name AS logged_by,
            dpl.created_at
        FROM daily_paper_logs dpl
        JOIN branches b ON b.id = dpl.branch_id
        LEFT JOIN users u ON u.id = dpl.logged_by
        WHERE {where}
        ORDER BY dpl.log_date DESC, b.code, dpl.paper_type
    """, tuple(params)) or []

    a4_total    = sum(r["total_sheets"] for r in rows if r["paper_type"] == "a4")
    b4_total    = sum(r["total_sheets"] for r in rows if r["paper_type"] == "b4")
    legal_total = sum(r["total_sheets"] for r in rows if r["paper_type"] == "legal")

    return {
        "rows":        rows,
        "a4_total":    a4_total,
        "b4_total":    b4_total,
        "legal_total": legal_total,
        "grand_total": a4_total + b4_total + legal_total,
    }

# ── Branch detail for a specific date ────────────────────────────────────────
@router.get("/prints/branch-detail/{branch_id}")
def get_branch_print_detail(
    branch_id: int,
    log_date: str = None,
    current_user: dict = Depends(require_nuwan)
):
    """
    Detailed print log for a specific branch on a given date.
    Returns each printer log with: who logged it, when, print counts per paper type.
    """
    from datetime import date, timedelta
    if not log_date:
        log_date = (date.today() - timedelta(days=1)).isoformat()

    rows = query("""
        SELECT
            pl.id              AS log_id,
            p.printer_code,
            p.model            AS printer_model,
            pl.print_count,
            pl.a4_single,
            pl.a4_double,
            pl.b4_single,
            pl.b4_double,
            pl.letter_single,
            pl.letter_double,
            pl.log_date,
            pl.created_at,
            u.full_name        AS logged_by_name,
            u.username         AS logged_by_user,
            b.code             AS branch_code,
            b.name             AS branch_name
        FROM print_logs pl
        JOIN printers p  ON p.id  = pl.printer_id
        JOIN branches b  ON b.id  = p.branch_id
        JOIN users u     ON u.id  = pl.logged_by
        WHERE b.id = %s
          AND pl.log_date = %s::date
        ORDER BY pl.created_at ASC
    """, (branch_id, log_date)) or []

    return {
        "branch_id":  branch_id,
        "log_date":   log_date,
        "logs":       rows,
        "total":      sum(r["print_count"] for r in rows),
    }


# ── Print Summary — filtered logs for Nuwan ─────────────────────────────────
@router.get("/prints/summary")
def get_print_summary(
    branch_id: int = None,
    date_from: str = None,
    date_to:   str = None,
    month:     int = None,
    year:      int = None,
    current_user: dict = Depends(require_nuwan)
):
    """
    Filtered print log summary for Nuwan.
    Returns per-row logs + branch totals + grand total.
    Supports filter by branch, date range, or month/year.
    """
    filters = ["1=1"]
    params  = []

    if branch_id:
        filters.append("b.id = %s")
        params.append(branch_id)

    if date_from:
        filters.append("pl.log_date >= %s::date")
        params.append(date_from)

    if date_to:
        filters.append("pl.log_date <= %s::date")
        params.append(date_to)

    if month and year:
        filters.append("EXTRACT(MONTH FROM pl.log_date) = %s")
        filters.append("EXTRACT(YEAR  FROM pl.log_date) = %s")
        params.extend([month, year])
    elif year:
        filters.append("EXTRACT(YEAR FROM pl.log_date) = %s")
        params.append(year)

    where = " AND ".join(filters)

    logs = query(f"""
        SELECT
            pl.id,
            pl.log_date,
            pl.print_count,
            pl.a4_single,  pl.a4_double,
            pl.b4_single,  pl.b4_double,
            pl.letter_single, pl.letter_double,
            pl.created_at,
            p.printer_code,
            p.model        AS printer_model,
            b.id           AS branch_id,
            b.code         AS branch_code,
            b.name         AS branch_name,
            u.full_name    AS logged_by,
            u.username     AS logged_by_user
        FROM print_logs pl
        JOIN printers p ON p.id  = pl.printer_id
        JOIN branches b ON b.id  = p.branch_id
        JOIN users    u ON u.id  = pl.logged_by
        WHERE {where}
        ORDER BY pl.log_date DESC, pl.created_at DESC, b.code, p.printer_code
        LIMIT 1000
    """, tuple(params)) or []

    # Branch totals
    branch_totals = query(f"""
        SELECT
            b.id           AS branch_id,
            b.code         AS branch_code,
            b.name         AS branch_name,
            COUNT(DISTINCT pl.log_date)                               AS days_logged,
            COUNT(pl.id)                                              AS log_count,
            COALESCE(SUM(pl.print_count),0)                           AS total_prints,
            COALESCE(SUM(pl.a4_single)+SUM(pl.a4_double),0)          AS a4_total,
            COALESCE(SUM(pl.b4_single)+SUM(pl.b4_double),0)          AS b4_total,
            COALESCE(SUM(pl.letter_single)+SUM(pl.letter_double),0)  AS legal_total
        FROM print_logs pl
        JOIN printers p ON p.id = pl.printer_id
        JOIN branches b ON b.id = p.branch_id
        WHERE {where}
        GROUP BY b.id, b.code, b.name
        ORDER BY total_prints DESC
    """, tuple(params)) or []

    grand_total   = sum(r["total_prints"] for r in branch_totals)
    grand_a4      = sum(r["a4_total"]     for r in branch_totals)
    grand_b4      = sum(r["b4_total"]     for r in branch_totals)
    grand_legal   = sum(r["legal_total"]  for r in branch_totals)

    return {
        "logs":          logs,
        "branch_totals": branch_totals,
        "grand_total":   grand_total,
        "grand_a4":      grand_a4,
        "grand_b4":      grand_b4,
        "grand_legal":   grand_legal,
        "record_count":  len(logs),
    }


# ── Print Summary Excel Export ───────────────────────────────────────────────
@router.get("/prints/summary/export")
def export_print_summary(
    branch_id: int = None,
    date_from: str = None,
    date_to:   str = None,
    month:     int = None,
    year:      int = None,
    current_user: dict = Depends(require_nuwan)
):
    """Export filtered print summary as Excel — for audit purposes."""
    from fastapi.responses import StreamingResponse
    import io
    from datetime import datetime as dt

    try:
        from openpyxl import Workbook # type: ignore
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side # type: ignore
        from openpyxl.utils import get_column_letter # type: ignore
    except ImportError:
        return {"error": "openpyxl not installed"}

    # Reuse summary logic
    filters = ["1=1"]
    params  = []
    if branch_id:
        filters.append("b.id = %s"); params.append(branch_id)
    if date_from:
        filters.append("pl.log_date >= %s::date"); params.append(date_from)
    if date_to:
        filters.append("pl.log_date <= %s::date"); params.append(date_to)
    if month and year:
        filters.append("EXTRACT(MONTH FROM pl.log_date) = %s")
        filters.append("EXTRACT(YEAR  FROM pl.log_date) = %s")
        params.extend([month, year])
    elif year:
        filters.append("EXTRACT(YEAR FROM pl.log_date) = %s"); params.append(year)

    where = " AND ".join(filters)

    logs = query(f"""
        SELECT pl.log_date, b.code AS branch_code, b.name AS branch_name,
               p.printer_code, p.model AS printer_model,
               pl.print_count,
               pl.a4_single, pl.a4_double,
               pl.b4_single, pl.b4_double,
               pl.letter_single, pl.letter_double,
               u.full_name AS logged_by, pl.created_at
        FROM print_logs pl
        JOIN printers p ON p.id = pl.printer_id
        JOIN branches b ON b.id = p.branch_id
        JOIN users    u ON u.id = pl.logged_by
        WHERE {where}
        ORDER BY pl.log_date DESC, b.code, p.printer_code
        LIMIT 5000
    """, tuple(params)) or []

    branch_totals = query(f"""
        SELECT b.code AS branch_code, b.name AS branch_name,
               COUNT(DISTINCT pl.log_date) AS days_logged,
               COALESCE(SUM(pl.print_count),0) AS total_prints,
               COALESCE(SUM(pl.a4_single)+SUM(pl.a4_double),0) AS a4_total,
               COALESCE(SUM(pl.b4_single)+SUM(pl.b4_double),0) AS b4_total,
               COALESCE(SUM(pl.letter_single)+SUM(pl.letter_double),0) AS legal_total
        FROM print_logs pl
        JOIN printers p ON p.id = pl.printer_id
        JOIN branches b ON b.id = p.branch_id
        WHERE {where}
        GROUP BY b.code, b.name ORDER BY total_prints DESC
    """, tuple(params)) or []

    # Build Excel
    wb  = Workbook()

    NAVY   = "1E3A5F"
    BLUE   = "0EA5E9"
    GREEN  = "10B981"
    WHITE  = "FFFFFF"
    ALTROW = "F8FAFC"
    BORDER = "D1D5DB"

    def hfont(sz=10, bold=True, col=WHITE):
        return Font(name="Arial", size=sz, bold=bold, color=col)
    def cfont(sz=9, bold=False, col="0F172A"):
        return Font(name="Arial", size=sz, bold=bold, color=col)
    def fill(c):
        return PatternFill("solid", start_color=c, fgColor=c)
    def thin():
        s = Side(style="thin", color=BORDER)
        return Border(left=s, right=s, top=s, bottom=s)
    def ctr():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def lft():
        return Alignment(horizontal="left", vertical="center")

    def fmt_date(v):
        if not v: return "—"
        try:
            if hasattr(v, 'strftime'): return v.strftime("%d/%m/%Y")
            return str(v)[:10]
        except: return str(v)

    def fmt_time(v):
        if not v: return "—"
        try:
            if hasattr(v, 'strftime'): return v.strftime("%d/%m/%Y %H:%M")
            from datetime import datetime
            return datetime.fromisoformat(str(v).replace("Z","")).strftime("%d/%m/%Y %H:%M")
        except: return str(v)

    # ── Sheet 1: Detailed Logs ────────────────────────────────
    ws1 = wb.active
    ws1.title = "Detailed Logs"

    ws1.merge_cells("A1:N1")
    ws1["A1"] = "SoftWave Print Management — Print Summary Report"
    ws1["A1"].font      = Font(name="Arial", size=14, bold=True, color=WHITE)
    ws1["A1"].fill      = fill(NAVY)
    ws1["A1"].alignment = ctr()
    ws1.row_dimensions[1].height = 28

    ws1.merge_cells("A2:N2")
    ws1["A2"] = f"Generated: {dt.now().strftime('%d %B %Y at %H:%M')}   |   Records: {len(logs)}"
    ws1["A2"].font      = Font(name="Arial", size=9, italic=True, color="64748B")
    ws1["A2"].fill      = fill("F1F5F9")
    ws1["A2"].alignment = ctr()
    ws1.row_dimensions[2].height = 16

    headers = [
        ("A3","Date"),("B3","Branch Code"),("C3","Branch Name"),
        ("D3","Printer Serial"),("E3","Model"),
        ("F3","Total Prints"),
        ("G3","A4 Single"),("H3","A4 Double"),
        ("I3","B4 Single"),("J3","B4 Double"),
        ("K3","Legal Single"),("L3","Legal Double"),
        ("M3","Logged By"),("N3","Logged At"),
    ]
    for addr, label in headers:
        c = ws1[addr]
        c.value     = label
        c.font      = Font(name="Arial", size=9, bold=True, color=WHITE)
        c.fill      = fill(BLUE)
        c.alignment = ctr()
        c.border    = thin()
    ws1.row_dimensions[3].height = 24

    for i, r in enumerate(logs):
        row = 4 + i
        bg  = WHITE if i % 2 == 0 else ALTROW
        vals = [
            fmt_date(r.get("log_date")),
            r.get("branch_code",""),
            r.get("branch_name",""),
            r.get("printer_code",""),
            r.get("printer_model",""),
            r.get("print_count",0),
            r.get("a4_single",0),  r.get("a4_double",0),
            r.get("b4_single",0),  r.get("b4_double",0),
            r.get("letter_single",0), r.get("letter_double",0),
            r.get("logged_by",""),
            fmt_time(r.get("created_at")),
        ]
        cols = "ABCDEFGHIJKLMN"
        for j, v in enumerate(vals):
            c = ws1[f"{cols[j]}{row}"]
            c.value     = v
            c.font      = cfont(bold=(j==5))
            c.fill      = fill(bg)
            c.alignment = lft() if j in [2,4,12] else ctr()
            c.border    = thin()
        ws1.row_dimensions[row].height = 16

    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 10
    ws1.column_dimensions["C"].width = 18
    ws1.column_dimensions["D"].width = 13
    ws1.column_dimensions["E"].width = 16
    ws1.column_dimensions["F"].width = 12
    for col in "GHIJKL":
        ws1.column_dimensions[col].width = 10
    ws1.column_dimensions["M"].width = 16
    ws1.column_dimensions["N"].width = 16
    ws1.freeze_panes = "A4"

    # ── Sheet 2: Branch Summary ───────────────────────────────
    ws2 = wb.create_sheet("Branch Summary")

    ws2.merge_cells("A1:G1")
    ws2["A1"] = "Branch Print Summary"
    ws2["A1"].font      = Font(name="Arial", size=13, bold=True, color=WHITE)
    ws2["A1"].fill      = fill(NAVY)
    ws2["A1"].alignment = ctr()
    ws2.row_dimensions[1].height = 26

    h2 = [("A2","Branch Code"),("B2","Branch Name"),("C2","Days Logged"),
          ("D2","Total Prints"),("E2","A4 Total"),("F2","B4 Total"),("G2","Legal Total")]
    for addr, label in h2:
        c = ws2[addr]
        c.value     = label
        c.font      = Font(name="Arial", size=9, bold=True, color=WHITE)
        c.fill      = fill(GREEN)
        c.alignment = ctr()
        c.border    = thin()
    ws2.row_dimensions[2].height = 22

    for i, r in enumerate(branch_totals):
        row = 3 + i
        bg  = WHITE if i % 2 == 0 else ALTROW
        vals = [r.get("branch_code",""), r.get("branch_name",""),
                r.get("days_logged",0),  r.get("total_prints",0),
                r.get("a4_total",0),     r.get("b4_total",0), r.get("legal_total",0)]
        for j, v in enumerate(vals):
            c = ws2[f"{'ABCDEFG'[j]}{row}"]
            c.value = v; c.font = cfont(bold=(j==3))
            c.fill = fill(bg); c.alignment = ctr() if j != 1 else lft()
            c.border = thin()
        ws2.row_dimensions[row].height = 16

    # Grand total row
    gr = 3 + len(branch_totals)
    grand = query(f"""
        SELECT COALESCE(SUM(pl.print_count),0) AS total,
               COALESCE(SUM(pl.a4_single)+SUM(pl.a4_double),0) AS a4,
               COALESCE(SUM(pl.b4_single)+SUM(pl.b4_double),0) AS b4,
               COALESCE(SUM(pl.letter_single)+SUM(pl.letter_double),0) AS legal
        FROM print_logs pl JOIN printers p ON p.id=pl.printer_id
        JOIN branches b ON b.id=p.branch_id WHERE {where}
    """, tuple(params), fetch="one") or {}

    ws2.merge_cells(f"A{gr}:C{gr}")
    ws2[f"A{gr}"] = "GRAND TOTAL"
    ws2[f"A{gr}"].font = Font(name="Arial", size=10, bold=True, color=WHITE)
    ws2[f"A{gr}"].fill = fill(NAVY)
    ws2[f"A{gr}"].alignment = ctr()
    for col, val in [("D", grand.get("total",0)), ("E", grand.get("a4",0)),
                     ("F", grand.get("b4",0)),     ("G", grand.get("legal",0))]:
        c = ws2[f"{col}{gr}"]
        c.value = val; c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        c.fill = fill(NAVY); c.alignment = ctr(); c.border = thin()

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 22
    for col in "CDEFG": ws2.column_dimensions[col].width = 14

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)

    fname = f"SoftWave_PrintSummary_{dt.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )



# ── Toner Audit Trail ────────────────────────────────────────────────────────
@router.get("/toner/audit")
def get_toner_audit(
    branch_id: int = None,
    current_user: dict = Depends(require_nuwan)
):
    """Full lifecycle of every toner installation — cost per copy included."""
    filters = ["1=1"]
    params  = []
    if branch_id:
        filters.append("b.id = %s")
        params.append(branch_id)

    where = " AND ".join(filters)

    rows = query(f"""
        SELECT
            ti.id              AS installation_id,
            p.printer_code,
            p.model            AS printer_model,
            b.id               AS branch_id,
            b.code             AS branch_code,
            b.name             AS branch_name,
            tm.model_code      AS toner_model,
            tm.brand,
            tm.yield_copies,
            COALESCE(tm.price_lkr, 0) AS price_lkr,
            ti.installed_at,
            u_inst.full_name   AS installed_by,
            -- Find next installation date for this printer (= replaced date)
            (SELECT ti2.installed_at FROM toner_installations ti2
             WHERE ti2.printer_id = ti.printer_id AND ti2.installed_at > ti.installed_at
             ORDER BY ti2.installed_at ASC LIMIT 1) AS replaced_at,
            -- Total prints logged during this installation
            COALESCE((
                SELECT SUM(pl.print_count) FROM print_logs pl
                WHERE pl.printer_id = ti.printer_id
                  AND pl.log_date >= ti.installed_at::date
                  AND (
                    pl.log_date < (
                        SELECT ti2.installed_at::date FROM toner_installations ti2
                        WHERE ti2.printer_id = ti.printer_id AND ti2.installed_at > ti.installed_at
                        ORDER BY ti2.installed_at ASC LIMIT 1
                    ) OR (
                        SELECT COUNT(*) FROM toner_installations ti2
                        WHERE ti2.printer_id = ti.printer_id AND ti2.installed_at > ti.installed_at
                    ) = 0
                  )
            ), 0) AS copies_made,
            ti.is_current
        FROM toner_installations ti
        JOIN printers p  ON p.id  = ti.printer_id
        JOIN branches b  ON b.id  = p.branch_id
        LEFT JOIN toner_models tm ON tm.id = ti.toner_model_id
        LEFT JOIN users u_inst ON u_inst.id = ti.installed_by
        WHERE {where}
        ORDER BY ti.installed_at DESC
        LIMIT 500
    """, tuple(params)) or []

    # Calculate cost per copy for each row
    result = []
    for r in rows:
        copies = int(r.get("copies_made") or 0)
        price  = float(r.get("price_lkr") or 0)
        cost_per_copy = round(price / copies, 2) if copies > 0 and price > 0 else None
        pct_used = round((copies / r["yield_copies"]) * 100, 1) if r.get("yield_copies") and copies > 0 else 0
        result.append({**r, "cost_per_copy": cost_per_copy, "pct_used": pct_used})

    return result


# ── Branch Performance Report ────────────────────────────────────────────────
@router.get("/reports/branch-performance")
def get_branch_performance(
    year:  int = None,
    month: int = None,
    current_user: dict = Depends(require_nuwan)
):
    """Monthly branch performance — prints, toner replacements, cost per copy."""
    from datetime import date as dt
    y = year  or dt.today().year
    m = month or dt.today().month

    branches = query("""
        SELECT
            b.id, b.code AS branch_code, b.name AS branch_name,
            -- Total prints this month
            COALESCE((
                SELECT SUM(pl.print_count) FROM print_logs pl
                JOIN printers p ON p.id = pl.printer_id
                WHERE p.branch_id = b.id
                  AND EXTRACT(YEAR  FROM pl.log_date) = %s
                  AND EXTRACT(MONTH FROM pl.log_date) = %s
            ), 0) AS total_prints,
            -- Days logged this month
            COALESCE((
                SELECT COUNT(DISTINCT pl.log_date) FROM print_logs pl
                JOIN printers p ON p.id = pl.printer_id
                WHERE p.branch_id = b.id
                  AND EXTRACT(YEAR  FROM pl.log_date) = %s
                  AND EXTRACT(MONTH FROM pl.log_date) = %s
            ), 0) AS days_logged,
            -- Toner replacements this month
            COALESCE((
                SELECT COUNT(*) FROM toner_installations ti
                JOIN printers p ON p.id = ti.printer_id
                WHERE p.branch_id = b.id
                  AND EXTRACT(YEAR  FROM ti.installed_at) = %s
                  AND EXTRACT(MONTH FROM ti.installed_at) = %s
            ), 0) AS toner_replacements,
            -- Toner cost this month
            COALESCE((
                SELECT SUM(COALESCE(tm.price_lkr, 0)) FROM toner_installations ti
                JOIN printers p ON p.id = ti.printer_id
                LEFT JOIN toner_models tm ON tm.id = ti.toner_model_id
                WHERE p.branch_id = b.id
                  AND EXTRACT(YEAR  FROM ti.installed_at) = %s
                  AND EXTRACT(MONTH FROM ti.installed_at) = %s
            ), 0) AS toner_cost_lkr,
            -- Active printers
            (SELECT COUNT(*) FROM printers p WHERE p.branch_id = b.id AND p.is_active = TRUE) AS printer_count
        FROM branches b
        WHERE b.is_active = TRUE
        ORDER BY total_prints DESC
    """, (y, m, y, m, y, m, y, m)) or []

    # Add cost per copy
    result = []
    for b in branches:
        tp   = int(b["total_prints"] or 0)
        cost = float(b["toner_cost_lkr"] or 0)
        result.append({
            **b,
            "cost_per_copy": round(cost / tp, 2) if tp > 0 and cost > 0 else None,
            "avg_daily_prints": round(tp / b["days_logged"], 0) if b["days_logged"] > 0 else 0,
        })

    grand_prints = sum(r["total_prints"] for r in result)
    grand_cost   = sum(float(r["toner_cost_lkr"]) for r in result)

    return {
        "year": y, "month": m,
        "branches": result,
        "grand_total_prints": grand_prints,
        "grand_toner_cost":   grand_cost,
        "grand_cost_per_copy": round(grand_cost / grand_prints, 2) if grand_prints > 0 and grand_cost > 0 else None,
    }



# ── Branch Performance Excel Export ─────────────────────────────────────────
@router.get("/reports/branch-performance/export")
def export_branch_performance(
    year:  int = None,
    month: int = None,
    current_user: dict = Depends(require_nuwan)
):
    from fastapi.responses import StreamingResponse
    from datetime import date as dt, datetime as dtt
    import io
    y = year  or dt.today().year
    m = month or dt.today().month

    try:
        from openpyxl import Workbook # type: ignore
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side # type: ignore
    except ImportError:
        return {"error": "openpyxl not installed"}

    # Get data
    branches = query("""
        SELECT b.id, b.code AS branch_code, b.name AS branch_name,
            COALESCE((SELECT SUM(pl.print_count) FROM print_logs pl JOIN printers p ON p.id=pl.printer_id
             WHERE p.branch_id=b.id AND EXTRACT(YEAR FROM pl.log_date)=%s AND EXTRACT(MONTH FROM pl.log_date)=%s),0) AS total_prints,
            COALESCE((SELECT COUNT(DISTINCT pl.log_date) FROM print_logs pl JOIN printers p ON p.id=pl.printer_id
             WHERE p.branch_id=b.id AND EXTRACT(YEAR FROM pl.log_date)=%s AND EXTRACT(MONTH FROM pl.log_date)=%s),0) AS days_logged,
            COALESCE((SELECT COUNT(*) FROM toner_installations ti JOIN printers p ON p.id=ti.printer_id
             WHERE p.branch_id=b.id AND EXTRACT(YEAR FROM ti.installed_at)=%s AND EXTRACT(MONTH FROM ti.installed_at)=%s),0) AS toner_replacements,
            COALESCE((SELECT SUM(COALESCE(tm.price_lkr,0)) FROM toner_installations ti JOIN printers p ON p.id=ti.printer_id
             LEFT JOIN toner_models tm ON tm.id=ti.toner_model_id
             WHERE p.branch_id=b.id AND EXTRACT(YEAR FROM ti.installed_at)=%s AND EXTRACT(MONTH FROM ti.installed_at)=%s),0) AS toner_cost_lkr,
            (SELECT COUNT(*) FROM printers p WHERE p.branch_id=b.id AND p.is_active=TRUE) AS printer_count
        FROM branches b WHERE b.is_active=TRUE ORDER BY total_prints DESC
    """, (y,m,y,m,y,m,y,m)) or []

    month_names = ["","January","February","March","April","May","June",
                   "July","August","September","October","November","December"]

    wb = Workbook()
    ws = wb.active
    ws.title = f"{month_names[m]} {y}"

    NAVY = "1E3A5F"; BLUE = "0EA5E9"; GREEN = "10B981"; WHITE = "FFFFFF"; ALTROW = "F8FAFC"
    def fill(c): return PatternFill("solid", start_color=c, fgColor=c)
    def border():
        s = Side(style="thin", color="D1D5DB")
        return Border(left=s, right=s, top=s, bottom=s)
    def ctr(): return Alignment(horizontal="center", vertical="center")
    def lft(): return Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A1:H1")
    ws["A1"] = f"SoftWave — Branch Performance Report: {month_names[m]} {y}"
    ws["A1"].font = Font(name="Arial", size=13, bold=True, color=WHITE)
    ws["A1"].fill = fill(NAVY); ws["A1"].alignment = ctr()
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Generated: {dtt.now().strftime('%d %B %Y at %H:%M')}"
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="64748B")
    ws["A2"].fill = fill("F1F5F9"); ws["A2"].alignment = ctr()

    headers = ["Branch Code","Branch Name","Total Prints","Days Logged","Avg Prints/Day",
               "Toner Replacements","Toner Cost (LKR)","Cost Per Copy (LKR)"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(3, j, h)
        c.font = Font(name="Arial", size=9, bold=True, color=WHITE)
        c.fill = fill(BLUE); c.alignment = ctr(); c.border = border()
    ws.row_dimensions[3].height = 22

    total_prints = 0; total_cost = 0
    for i, b in enumerate(branches):
        row = 4 + i
        bg = WHITE if i%2==0 else ALTROW
        tp = int(b["total_prints"] or 0)
        dl = int(b["days_logged"] or 0)
        cost = float(b["toner_cost_lkr"] or 0)
        avg_day = round(tp/dl, 0) if dl > 0 else 0
        cpc = round(cost/tp, 2) if tp > 0 and cost > 0 else ""
        vals = [b["branch_code"], b["branch_name"], tp, dl, avg_day,
                int(b["toner_replacements"] or 0), round(cost, 2), cpc]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row, j, v)
            c.font = Font(name="Arial", size=9, bold=(j==3))
            c.fill = fill(bg); c.alignment = ctr() if j != 2 else lft()
            c.border = border()
        total_prints += tp; total_cost += cost

    # Grand total row
    gr = 4 + len(branches)
    ws.merge_cells(f"A{gr}:B{gr}")
    ws[f"A{gr}"] = "GRAND TOTAL"
    ws[f"A{gr}"].font = Font(name="Arial", size=10, bold=True, color=WHITE)
    ws[f"A{gr}"].fill = fill(NAVY); ws[f"A{gr}"].alignment = ctr()
    for col, val in [(3, total_prints), (7, round(total_cost, 2)),
                     (8, round(total_cost/total_prints, 2) if total_prints > 0 and total_cost > 0 else "")]:
        c = ws.cell(gr, col, val)
        c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        c.fill = fill(NAVY); c.alignment = ctr(); c.border = border()

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    for col in "CDEFGH": ws.column_dimensions[col].width = 14
    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"BranchReport_{y}_{str(m).zfill(2)}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})

# ── Monthly print totals ─────────────────────────────────────────────────────
@router.get("/prints/monthly")
def get_monthly_prints(
    year:  int = Query(default=None),
    month: int = Query(default=None),
    current_user: dict = Depends(require_nuwan)
):
    today = date.today()
    y = year  or today.year
    m = month or today.month

    rows = query("""
        SELECT
            b.id            AS branch_id,
            b.code          AS branch_code,
            b.name          AS branch_name,
            pl.log_date,
            COALESCE(SUM(pl.print_count), 0) AS daily_total
        FROM branches b
        JOIN printers p  ON p.branch_id = b.id AND p.is_active = TRUE
        JOIN print_logs pl ON pl.printer_id = p.id
            AND EXTRACT(YEAR  FROM pl.log_date) = %s
            AND EXTRACT(MONTH FROM pl.log_date) = %s
        WHERE b.is_active = TRUE
        GROUP BY b.id, b.code, b.name, pl.log_date
        ORDER BY b.code, pl.log_date
    """, (y, m))

    # Aggregate per branch
    branch_map = {}
    for r in (rows or []):
        bid = r["branch_id"]
        if bid not in branch_map:
            branch_map[bid] = {
                "branch_id":   bid,
                "branch_code": r["branch_code"],
                "branch_name": r["branch_name"],
                "total":       0,
                "days":        {},
            }
        day_str = r["log_date"].isoformat() if hasattr(r["log_date"], "isoformat") else str(r["log_date"])
        branch_map[bid]["days"][day_str] = int(r["daily_total"])
        branch_map[bid]["total"] += int(r["daily_total"])

    return {
        "year":        y,
        "month":       m,
        "grand_total": sum(v["total"] for v in branch_map.values()),
        "branches":    list(branch_map.values()),
    }


# ── Monthly Excel export data ────────────────────────────────────────────────
@router.get("/prints/export")
def get_export_data(
    year:  int = Query(default=None),
    month: int = Query(default=None),
    current_user: dict = Depends(require_nuwan)
):
    today = date.today()
    y = year  or today.year
    m = month or today.month

    rows = query("""
        SELECT
            b.code          AS branch_code,
            b.name          AS branch_name,
            p.printer_code,
            p.model,
            pl.log_date,
            pl.print_count,
            u.full_name     AS logged_by
        FROM branches b
        JOIN printers p  ON p.branch_id = b.id AND p.is_active = TRUE
        JOIN print_logs pl ON pl.printer_id = p.id
            AND EXTRACT(YEAR  FROM pl.log_date) = %s
            AND EXTRACT(MONTH FROM pl.log_date) = %s
        LEFT JOIN users u ON u.id = pl.logged_by
        WHERE b.is_active = TRUE
        ORDER BY b.code, p.printer_code, pl.log_date
    """, (y, m))

    return {
        "year":   y,
        "month":  m,
        "rows":   rows or [],
    }


# ── Toner models (for request modal) ────────────────────────────────────────
@router.get("/toner-models")
def get_toner_models_for_request(current_user: dict = Depends(require_nuwan)):
    return query("SELECT id, model_code, brand, yield_copies FROM toner_models ORDER BY model_code") or []


# ── Submit toner request from Nuwan dashboard ────────────────────────────────
class NuwanTonerRequestBody(BaseModel):
    printer_id:     int
    toner_model_id: int
    priority:       str = "urgent"
    notes:          str = ""


@router.post("/request-toner")
def nuwan_request_toner(body: NuwanTonerRequestBody, current_user: dict = Depends(require_nuwan)):
    if body.priority not in ("normal", "urgent", "critical"):
        raise HTTPException(status_code=400, detail="Invalid priority")
    req = query(
        "INSERT INTO replacement_requests "
        "(request_type, printer_id, toner_model_id, quantity, priority, notes, requested_by) "
        "VALUES ('toner', %s, %s, 1, %s, %s, %s) RETURNING id",
        (body.printer_id, body.toner_model_id, body.priority, body.notes, int(current_user["sub"])),
        fetch="one"
    )
    return {"message": "Toner request submitted", "id": req["id"]}


# ── Nuwan's own requests ─────────────────────────────────────────────────────
@router.get("/my-requests")
def get_nuwan_requests(current_user: dict = Depends(require_nuwan)):
    return query(
        "SELECT rr.*, p.printer_code, b.code AS branch_code, b.name AS branch_name, "
        "tm.model_code AS toner_model_code, "
        "u.full_name AS requested_by_name, rv.full_name AS reviewed_by_name, "
        "ds.full_name AS dispatched_by_name "
        "FROM replacement_requests rr "
        "JOIN printers p  ON p.id  = rr.printer_id "
        "JOIN branches b  ON b.id  = p.branch_id "
        "LEFT JOIN toner_models tm ON tm.id  = rr.toner_model_id "
        "LEFT JOIN users u  ON u.id  = rr.requested_by "
        "LEFT JOIN users rv ON rv.id = rr.reviewed_by "
        "LEFT JOIN users ds ON ds.id = rr.dispatched_by "
        "WHERE rr.requested_by = %s AND rr.request_type = 'toner' "
        "ORDER BY rr.requested_at DESC LIMIT 50",
        (int(current_user["sub"]),)
    ) or []