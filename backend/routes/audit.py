"""
Audit routes — Toner, Paper, Hardware audit data + Cost per copy.
Add to backend/routes/audit.py and register in main.py
"""

from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
from db.database import query
from middleware.auth import get_current_user, require_role
from datetime import datetime

router = APIRouter(prefix="/api/audit", tags=["Audit"])

def require_manager(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ("manager", "dba", "nuwan"):
        raise HTTPException(status_code=403, detail="Access denied")
    return current_user


# ── TONER AUDIT ──────────────────────────────────────────────
@router.get("/toner")
def get_toner_audit(
    branch_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to:   Optional[str] = None,
    current_user: dict = Depends(require_manager)
):
    filters = ["1=1"]
    params  = []
    if branch_id:
        filters.append("b.id = %s"); params.append(branch_id)
    if date_from:
        filters.append("ti.installed_at::date >= %s::date"); params.append(date_from)
    if date_to:
        filters.append("ti.installed_at::date <= %s::date"); params.append(date_to)
    where = " AND ".join(filters)

    rows = query(f"""
        SELECT
            ti.id              AS installation_id,
            b.id               AS branch_id,
            b.code             AS branch_code,
            b.name             AS branch_name,
            p.printer_code,
            p.model            AS printer_model,
            tm.model_code      AS toner_model,
            tm.brand,
            COALESCE(tm.yield_copies, 0)   AS yield_copies,
            COALESCE(tm.price_lkr, 0)      AS price_lkr,
            ti.installed_at,
            ti.is_current,
            COALESCE(ti.total_copies_done, 0) AS copies_made,
            -- Replaced date = next installation for same printer
            (SELECT ti2.installed_at FROM toner_installations ti2
             WHERE ti2.printer_id = ti.printer_id
               AND ti2.installed_at > ti.installed_at
             ORDER BY ti2.installed_at LIMIT 1) AS replaced_at,
            -- Remaining copies
            GREATEST(0, COALESCE(ti.yield_copies,0) - COALESCE(ti.total_copies_done,0))
                AS copies_remaining
        FROM toner_installations ti
        JOIN printers p ON p.id = ti.printer_id
        JOIN branches b ON b.id = p.branch_id
        LEFT JOIN toner_models tm ON tm.id = ti.toner_model_id
        WHERE {where}
        ORDER BY ti.installed_at DESC
        LIMIT 500
    """, tuple(params)) or []

    result = []
    for r in rows:
        copies    = int(r.get("copies_made") or 0)
        yield_c   = int(r.get("yield_copies") or 1)
        price     = float(r.get("price_lkr") or 0)
        pct_used  = round((copies / yield_c) * 100, 1) if yield_c > 0 else 0
        cpc       = round(price / copies, 2) if copies > 0 and price > 0 else None
        result.append({**r, "pct_used": pct_used, "cost_per_copy": cpc})

    return result


# ── PAPER AUDIT ──────────────────────────────────────────────
@router.get("/paper")
def get_paper_audit(
    branch_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to:   Optional[str] = None,
    current_user: dict = Depends(require_manager)
):
    filters = ["1=1"]
    params  = []
    if branch_id:
        filters.append("b.id = %s"); params.append(branch_id)
    if date_from:
        filters.append("pm.created_at::date >= %s::date"); params.append(date_from)
    if date_to:
        filters.append("pm.created_at::date <= %s::date"); params.append(date_to)
    where = " AND ".join(filters)

    # Get dispatches (received by branch)
    dispatches = query(f"""
        SELECT
            pm.id,
            pm.created_at        AS received_date,
            b.id                 AS branch_id,
            b.code               AS branch_code,
            b.name               AS branch_name,
            pt.id                AS paper_type_id,
            pt.name              AS paper_name,
            pt.size,
            pt.gsm,
            COALESCE(pt.price_per_ream, 0) AS price_per_ream,
            pm.quantity          AS reams_received,
            -- sheets per ream = 500
            pm.quantity * 500    AS sheets_received,
            u.full_name          AS dispatched_by
        FROM paper_movements pm
        JOIN branches b  ON b.id  = pm.branch_id
        JOIN paper_types pt ON pt.id = pm.paper_type_id
        LEFT JOIN users u ON u.id = pm.performed_by
        WHERE pm.movement_type = 'OUT'
          AND {where}
        ORDER BY pm.created_at DESC
        LIMIT 500
    """, tuple(params)) or []

    # For each dispatch, get usage from daily_paper_logs for that branch/type
    result = []
    for d in dispatches:
        bid   = d["branch_id"]
        pt_id = d["paper_type_id"]
        size  = (d["size"] or "").lower()
        # Map size to paper_type key
        type_key = "legal" if size == "letter" else size.lower()

        usage = query("""
            SELECT
                COALESCE(SUM(single_side), 0) AS total_single,
                COALESCE(SUM(double_side), 0) AS total_double,
                COALESCE(SUM(CASE WHEN paper_type='a4'    THEN waste_a4    ELSE 0 END), 0) AS waste_a4,
                COALESCE(SUM(CASE WHEN paper_type='b4'    THEN waste_b4    ELSE 0 END), 0) AS waste_b4,
                COALESCE(SUM(CASE WHEN paper_type='legal' THEN waste_legal ELSE 0 END), 0) AS waste_legal
            FROM daily_paper_logs
            WHERE branch_id = %s AND paper_type = %s
        """, (bid, type_key), fetch="one") or {}

        total_single = int(usage.get("total_single") or 0)
        total_double = int(usage.get("total_double") or 0)
        waste_val    = int(usage.get(f"waste_{type_key}") or 0)
        total_used   = total_single + total_double
        total_sheets = int(d.get("sheets_received") or 0)
        remaining    = max(0, total_sheets - total_used - waste_val)
        price_ream   = float(d.get("price_per_ream") or 0)
        total_cost   = round(price_ream * int(d.get("reams_received") or 0), 2)
        waste_cost   = round((waste_val / 500) * price_ream, 2) if price_ream > 0 else 0

        result.append({
            **d,
            "total_single":   total_single,
            "total_double":   total_double,
            "total_used":     total_used,
            "waste_sheets":   waste_val,
            "remaining":      remaining,
            "total_cost_lkr": total_cost,
            "waste_cost_lkr": waste_cost,
        })

    return result


# ── HARDWARE AUDIT ───────────────────────────────────────────
@router.get("/hardware")
def get_hardware_audit(
    branch_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to:   Optional[str] = None,
    current_user: dict = Depends(require_manager)
):
    filters = ["hr.status = 'installed'"]
    params  = []
    if branch_id:
        filters.append("b.id = %s"); params.append(branch_id)
    if date_from:
        filters.append("hr.installed_at::date >= %s::date"); params.append(date_from)
    if date_to:
        filters.append("hr.installed_at::date <= %s::date"); params.append(date_to)
    where = " AND ".join(filters)

    rows = query(f"""
        SELECT
            hr.id,
            b.id               AS branch_id,
            b.code             AS branch_code,
            b.name             AS branch_name,
            p.printer_code,
            p.model            AS printer_model,
            hr.part_name,
            hr.installed_at,
            hr.price_lkr,
            hr.estimated_life_months,
            -- Estimated next replacement date
            hr.installed_at + (hr.estimated_life_months || ' months')::INTERVAL
                AS estimated_next_date,
            -- Days until next replacement
            EXTRACT(DAY FROM
                (hr.installed_at + (hr.estimated_life_months || ' months')::INTERVAL)
                - NOW()
            )::INTEGER AS days_remaining,
            u_inst.full_name   AS installed_by_name,
            u_req.full_name    AS requested_by_name
        FROM hardware_requests hr
        JOIN branches b ON b.id = hr.branch_id
        JOIN printers p ON p.id = hr.printer_id
        LEFT JOIN users u_inst ON u_inst.id = hr.installed_by
        LEFT JOIN users u_req  ON u_req.id  = hr.requested_by
        WHERE {where}
        ORDER BY hr.installed_at DESC
        LIMIT 500
    """, tuple(params)) or []

    result = []
    for r in rows:
        days = r.get("days_remaining")
        if days is None:
            status = "unknown"
        elif days <= 0:
            status = "overdue"
        elif days <= 30:
            status = "due_soon"
        elif days <= 90:
            status = "warning"
        else:
            status = "good"
        result.append({**r, "life_status": status})

    return result


# ── PERFORMANCE SUMMARY (Cost per copy) ─────────────────────
@router.get("/performance")
def get_performance_summary(
    branch_id: Optional[int] = None,
    month:     Optional[int] = None,
    year:      Optional[int] = None,
    current_user: dict = Depends(require_manager)
):
    from datetime import date
    y = year  or date.today().year
    m = month or date.today().month

    branches = query("""
        SELECT id, code AS branch_code, name AS branch_name
        FROM branches WHERE is_active = TRUE ORDER BY name
    """) or []

    result = []
    for b in branches:
        bid = b["id"]

        # Total prints this month
        prints = query("""
            SELECT COALESCE(SUM(pl.print_count), 0) AS total
            FROM print_logs pl
            JOIN printers p ON p.id = pl.printer_id
            WHERE p.branch_id = %s
              AND EXTRACT(YEAR FROM pl.log_date) = %s
              AND EXTRACT(MONTH FROM pl.log_date) = %s
        """, (bid, y, m), fetch="one") or {}

        # Toner cost this month
        toner = query("""
            SELECT
                COUNT(*) AS replacements,
                COALESCE(SUM(COALESCE(tm.price_lkr, 0)), 0) AS cost
            FROM toner_installations ti
            JOIN printers p ON p.id = ti.printer_id
            LEFT JOIN toner_models tm ON tm.id = ti.toner_model_id
            WHERE p.branch_id = %s
              AND EXTRACT(YEAR FROM ti.installed_at) = %s
              AND EXTRACT(MONTH FROM ti.installed_at) = %s
        """, (bid, y, m), fetch="one") or {}

        # Paper cost this month
        paper = query("""
            SELECT COALESCE(SUM(pm.quantity * COALESCE(pt.price_per_ream, 0)), 0) AS cost,
                   COALESCE(SUM(pm.quantity), 0) AS reams
            FROM paper_movements pm
            JOIN paper_types pt ON pt.id = pm.paper_type_id
            WHERE pm.branch_id = %s
              AND pm.movement_type = 'OUT'
              AND EXTRACT(YEAR FROM pm.created_at) = %s
              AND EXTRACT(MONTH FROM pm.created_at) = %s
        """, (bid, y, m), fetch="one") or {}

        # Hardware cost this month
        hardware = query("""
            SELECT COALESCE(SUM(COALESCE(price_lkr, 0)), 0) AS cost,
                   COUNT(*) AS installs
            FROM hardware_requests
            WHERE branch_id = %s AND status = 'installed'
              AND EXTRACT(YEAR FROM installed_at) = %s
              AND EXTRACT(MONTH FROM installed_at) = %s
        """, (bid, y, m), fetch="one") or {}

        # Per-printer breakdown
        printers = query("""
            SELECT
                p.id           AS printer_id,
                p.printer_code,
                p.model,
                COALESCE((
                    SELECT SUM(pl.print_count) FROM print_logs pl
                    WHERE pl.printer_id = p.id
                      AND EXTRACT(YEAR FROM pl.log_date) = %s
                      AND EXTRACT(MONTH FROM pl.log_date) = %s
                ), 0) AS prints,
                COALESCE((
                    SELECT SUM(COALESCE(tm.price_lkr, 0))
                    FROM toner_installations ti
                    LEFT JOIN toner_models tm ON tm.id = ti.toner_model_id
                    WHERE ti.printer_id = p.id
                      AND EXTRACT(YEAR FROM ti.installed_at) = %s
                      AND EXTRACT(MONTH FROM ti.installed_at) = %s
                ), 0) AS toner_cost,
                COALESCE((
                    SELECT SUM(COALESCE(price_lkr, 0))
                    FROM hardware_requests hr
                    WHERE hr.printer_id = p.id AND hr.status = 'installed'
                      AND EXTRACT(YEAR FROM hr.installed_at) = %s
                      AND EXTRACT(MONTH FROM hr.installed_at) = %s
                ), 0) AS hardware_cost
            FROM printers p
            WHERE p.branch_id = %s AND p.is_active = TRUE
            ORDER BY p.printer_code
        """, (y, m, y, m, y, m, bid)) or []

        total_prints   = int(prints.get("total") or 0)
        toner_cost     = float(toner.get("cost") or 0)
        paper_cost     = float(paper.get("cost") or 0)
        hardware_cost  = float(hardware.get("cost") or 0)
        total_cost     = toner_cost + paper_cost + hardware_cost
        total_reams    = int(paper.get("reams") or 0)

        # Paper cost per branch = total paper dispatched / total prints
        # (allocate proportionally since paper isn't tracked per printer)
        for pr in printers:
            pr_prints = int(pr.get("prints") or 0)
            pr_toner  = float(pr.get("toner_cost") or 0)
            pr_hw     = float(pr.get("hardware_cost") or 0)
            # Allocate paper cost proportionally by print share
            pr_paper  = round(
                paper_cost * (pr_prints / total_prints), 2
            ) if total_prints > 0 else 0
            pr_total  = pr_toner + pr_hw + pr_paper
            pr["paper_cost"]   = pr_paper
            pr["total_cost"]   = pr_total
            pr["cost_per_copy"] = round(pr_total / pr_prints, 4) if pr_prints > 0 else None

        result.append({
            "branch_id":       bid,
            "branch_code":     b["branch_code"],
            "branch_name":     b["branch_name"],
            "total_prints":    total_prints,
            "toner_cost":      toner_cost,
            "paper_cost":      paper_cost,
            "hardware_cost":   hardware_cost,
            "total_cost":      total_cost,
            "toner_replacements": int(toner.get("replacements") or 0),
            "hardware_installs":  int(hardware.get("installs") or 0),
            "paper_reams":     total_reams,
            "cost_per_copy":   round(total_cost / total_prints, 4) if total_prints > 0 else None,
            "printers":        printers,
        })

    grand_prints = sum(r["total_prints"] for r in result)
    grand_cost   = sum(r["total_cost"]   for r in result)

    return {
        "year": y, "month": m,
        "branches": result,
        "grand_prints":     grand_prints,
        "grand_toner_cost": sum(r["toner_cost"]    for r in result),
        "grand_paper_cost": sum(r["paper_cost"]    for r in result),
        "grand_hw_cost":    sum(r["hardware_cost"] for r in result),
        "grand_total_cost": grand_cost,
        "grand_cpc":        round(grand_cost / grand_prints, 4) if grand_prints > 0 else None,
    }


# ── PERFORMANCE EXCEL EXPORT ─────────────────────────────────
@router.get("/performance/export")
def export_performance(
    branch_id: Optional[int] = None,
    month:     Optional[int] = None,
    year:      Optional[int] = None,
    current_user: dict = Depends(require_manager)
):
    from datetime import date as dt_date
    import io
    y = year  or dt_date.today().year
    m = month or dt_date.today().month

    try:
        from openpyxl import Workbook # type: ignore
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side # type: ignore
    except ImportError:
        return {"error": "openpyxl not installed"}

    # Re-use performance endpoint logic
    data = get_performance_summary(branch_id=branch_id, month=m, year=y, current_user=current_user)

    NAVY="1E3A5F"; BLUE="0EA5E9"; GREEN="10B981"; WHITE="FFFFFF"; ALTROW="F8FAFC"
    def fill(c): return PatternFill("solid", start_color=c, fgColor=c)
    def bdr():
        s=Side(style="thin",color="D1D5DB")
        return Border(left=s,right=s,top=s,bottom=s)
    def ctr(): return Alignment(horizontal="center",vertical="center")
    def lft(): return Alignment(horizontal="left",vertical="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "Performance Summary"

    month_names=["","Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

    ws.merge_cells("A1:I1")
    ws["A1"] = f"SoftWave — Performance Summary: {month_names[m]} {y}"
    ws["A1"].font=Font(name="Arial",size=13,bold=True,color=WHITE)
    ws["A1"].fill=fill(NAVY); ws["A1"].alignment=ctr()
    ws.row_dimensions[1].height=26

    headers=["Branch","Total Prints","Toner Cost (LKR)","Paper Cost (LKR)",
             "Hardware Cost (LKR)","Total Cost (LKR)","Cost/Copy (LKR)",
             "Toner Replacements","Paper Reams"]
    for j,h in enumerate(headers,1):
        c=ws.cell(3,j,h)
        c.font=Font(name="Arial",size=9,bold=True,color=WHITE)
        c.fill=fill(BLUE); c.alignment=ctr(); c.border=bdr()

    for i,b in enumerate(data["branches"]):
        row=4+i; bg=WHITE if i%2==0 else ALTROW
        cpc=b["cost_per_copy"]
        vals=[b["branch_name"],b["total_prints"],round(b["toner_cost"],2),
              round(b["paper_cost"],2),round(b["hardware_cost"],2),
              round(b["total_cost"],2),
              round(cpc,4) if cpc else "",
              b["toner_replacements"],b["paper_reams"]]
        for j,v in enumerate(vals,1):
            c=ws.cell(row,j,v)
            c.font=Font(name="Arial",size=9,bold=(j==2))
            c.fill=fill(bg); c.alignment=lft() if j==1 else ctr(); c.border=bdr()

    # Grand total
    gr=4+len(data["branches"])
    ws.cell(gr,1,"GRAND TOTAL").font=Font(name="Arial",size=10,bold=True,color=WHITE)
    ws.cell(gr,1).fill=fill(NAVY); ws.cell(gr,1).alignment=ctr()
    for col,val in [(2,data["grand_prints"]),(3,round(data["grand_toner_cost"],2)),
                    (4,round(data["grand_paper_cost"],2)),(5,round(data["grand_hw_cost"],2)),
                    (6,round(data["grand_total_cost"],2)),
                    (7,round(data["grand_cpc"],4) if data["grand_cpc"] else "")]:
        c=ws.cell(gr,col,val)
        c.font=Font(name="Arial",size=10,bold=True,color=WHITE)
        c.fill=fill(NAVY); c.alignment=ctr(); c.border=bdr()

    ws.column_dimensions["A"].width=22
    for col in "BCDEFGHI": ws.column_dimensions[col].width=16
    ws.freeze_panes="A4"

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    fname=f"SoftWave_Performance_{y}_{str(m).zfill(2)}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":f"attachment; filename={fname}"})


# ── Branches for audit filters ────────────────────────────────
@router.get("/branches")
def get_audit_branches(current_user: dict = Depends(require_manager)):
    return query("SELECT id, code, name FROM branches WHERE is_active=TRUE ORDER BY name") or []
