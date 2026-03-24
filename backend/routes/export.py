"""
SoftWave — Toner Audit Export
Generates a professional Excel report for toner replacement audit.
Manager/DBA only.
"""

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from middleware.auth import require_role
from db.database import query
import io
from datetime import datetime

router = APIRouter(prefix="/api/export", tags=["Export"])

_allowed = require_role("manager", "dba")


@router.get("/toner-audit")
def export_toner_audit(current_user: dict = Depends(_allowed)):
    """
    Download a full toner replacement audit Excel report.
    Columns: Branch | Printer | Toner Model |
             Requested By & Date | Approved By & Date |
             Dispatched By & Date | Replaced By & Date | Status
    """
    try:
        from openpyxl import Workbook # type: ignore
        from openpyxl.styles import ( # type: ignore
            Font, PatternFill, Alignment, Border, Side, GradientFill
        )
        from openpyxl.utils import get_column_letter # type: ignore
    except ImportError:
        return {"error": "openpyxl not installed"}

    # ── Fetch full audit data ─────────────────────────────────
    rows = query("""
        SELECT
            rr.id                           AS request_id,
            b.code                          AS branch_code,
            b.name                          AS branch_name,
            p.printer_code,
            p.model                         AS printer_model,
            tm.model_code                   AS toner_model,

            -- Requested
            u_req.full_name                 AS requested_by,
            rr.requested_at,

            -- Approved / Rejected
            u_rev.full_name                 AS reviewed_by,
            rr.reviewed_at,
            rr.review_note,

            -- Dispatched
            u_dis.full_name                 AS dispatched_by,
            rr.dispatched_at,
            rr.dispatch_note,

            -- Replaced (toner_installations tracks who installed & when)
            u_ins.full_name                 AS replaced_by,
            ti.installed_at                 AS replaced_at,

            rr.status,
            rr.priority,
            rr.quantity,
            rr.notes                        AS request_notes

        FROM replacement_requests rr
        JOIN printers  p    ON p.id   = rr.printer_id
        JOIN branches  b    ON b.id   = p.branch_id
        LEFT JOIN toner_models tm ON tm.id = rr.toner_model_id
        LEFT JOIN users u_req     ON u_req.id = rr.requested_by
        LEFT JOIN users u_rev     ON u_rev.id = rr.reviewed_by
        LEFT JOIN users u_dis     ON u_dis.id = rr.dispatched_by
        -- Link to toner installation: find the installation done after dispatch
        LEFT JOIN toner_installations ti
               ON ti.printer_id   = p.id
              AND ti.is_current   = TRUE
        LEFT JOIN users u_ins     ON u_ins.id = ti.installed_by

        WHERE rr.request_type = 'toner'
        ORDER BY rr.requested_at DESC
    """) or []

    # ── Build Excel ───────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Toner Audit"

    # Colours
    C_HEADER_BG  = "1E3A5F"   # dark navy
    C_HEADER_FG  = "FFFFFF"
    C_SEC1_BG    = "EBF4FF"   # light blue  — request
    C_SEC2_BG    = "EEFBF0"   # light green — approved
    C_SEC3_BG    = "FFF8EC"   # light amber — dispatched
    C_SEC4_BG    = "F3EEFF"   # light purple— replaced
    C_ALT_ROW    = "F8FAFC"
    C_WHITE      = "FFFFFF"
    C_BORDER     = "D1D5DB"

    def hdr_font(size=10, bold=True, color=C_HEADER_FG):
        return Font(name="Arial", size=size, bold=bold, color=color)

    def cell_font(size=9, bold=False, color="0F172A"):
        return Font(name="Arial", size=size, bold=bold, color=color)

    def fill(hex_color):
        return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

    def thin_border():
        s = Side(style="thin", color=C_BORDER)
        return Border(left=s, right=s, top=s, bottom=s)

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    def fmt_dt(dt, include_time=True):
        if not dt:
            return "—"
        try:
            if isinstance(dt, str):
                dt = datetime.fromisoformat(dt.replace("Z",""))
            if include_time:
                return dt.strftime("%d/%m/%Y  %H:%M")
            return dt.strftime("%d/%m/%Y")
        except Exception:
            return str(dt)

    # ── Row 1: Report title ───────────────────────────────────
    ws.merge_cells("A1:T1")
    ws["A1"] = "SoftWave Print Management — Toner Replacement Audit Report"
    ws["A1"].font      = Font(name="Arial", size=14, bold=True, color=C_HEADER_FG)
    ws["A1"].fill      = fill(C_HEADER_BG)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 30

    # ── Row 2: Generated date ─────────────────────────────────
    ws.merge_cells("A2:T2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%d %B %Y at %H:%M')}   |   Total Records: {len(rows)}"
    ws["A2"].font      = Font(name="Arial", size=9, italic=True, color="64748B")
    ws["A2"].fill      = fill("F1F5F9")
    ws["A2"].alignment = center()
    ws.row_dimensions[2].height = 18

    # ── Row 3: blank spacer ────────────────────────────────────
    ws.row_dimensions[3].height = 6

    # ── Row 4: Section labels ──────────────────────────────────
    sec_labels = [
        ("A4:D4", "PRINTER DETAILS",   C_HEADER_BG),
        ("E4:G4", "REQUEST",           "1D6FA4"),
        ("H4:J4", "APPROVED / REJECTED","1A7A4A"),
        ("K4:M4", "DISPATCHED",        "A05C00"),
        ("N4:P4", "REPLACED (INSTALLED)","5B21B6"),
        ("Q4:T4", "STATUS & NOTES",    "374151"),
    ]
    for cell_range, label, bg in sec_labels:
        ws.merge_cells(cell_range)
        c = ws[cell_range.split(":")[0]]
        c.value     = label
        c.font      = hdr_font(9)
        c.fill      = fill(bg)
        c.alignment = center()
    ws.row_dimensions[4].height = 20

    # ── Row 5: Column headers ─────────────────────────────────
    headers = [
        # Printer
        ("A5","Branch Code"),("B5","Branch Name"),
        ("C5","Printer Serial"),("D5","Toner Model"),
        # Request
        ("E5","Requested By"),("F5","Requested Date"),("G5","Request Notes"),
        # Approved
        ("H5","Approved/Rejected By"),("I5","Approved Date"),("J5","Review Note"),
        # Dispatched
        ("K5","Dispatched By"),("L5","Dispatched Date"),("M5","Dispatch Note"),
        # Replaced
        ("N5","Installed By"),("O5","Installed Date"),("P5","Qty"),
        # Status
        ("Q5","Status"),("R5","Priority"),("S5","Req ID"),("T5",""),
    ]
    sec_bg = {
        "A5":"D9E8F5","B5":"D9E8F5","C5":"D9E8F5","D5":"D9E8F5",
        "E5":"CCEEDD","F5":"CCEEDD","G5":"CCEEDD",
        "H5":"D4EDDA","I5":"D4EDDA","J5":"D4EDDA",
        "K5":"FFF0CC","L5":"FFF0CC","M5":"FFF0CC",
        "N5":"E9D5FF","O5":"E9D5FF","P5":"E9D5FF",
        "Q5":"E5E7EB","R5":"E5E7EB","S5":"E5E7EB","T5":"E5E7EB",
    }
    for addr, label in headers:
        c = ws[addr]
        c.value     = label
        c.font      = Font(name="Arial", size=9, bold=True, color="0F172A")
        c.fill      = fill(sec_bg.get(addr, "E5E7EB"))
        c.alignment = center()
        c.border    = thin_border()
    ws.row_dimensions[5].height = 30

    # ── Data rows ─────────────────────────────────────────────
    status_colors = {
        "dispatched": "D1FAE5",
        "approved":   "DBEAFE",
        "pending":    "FEF9C3",
        "rejected":   "FEE2E2",
    }

    for i, r in enumerate(rows):
        row_num = 6 + i
        bg = status_colors.get(r.get("status",""), C_WHITE if i%2==0 else C_ALT_ROW)

        vals = [
            r.get("branch_code",""),
            r.get("branch_name",""),
            r.get("printer_code",""),
            r.get("toner_model",""),
            r.get("requested_by","—"),
            fmt_dt(r.get("requested_at")),
            r.get("request_notes",""),
            r.get("reviewed_by","—"),
            fmt_dt(r.get("reviewed_at")),
            r.get("review_note",""),
            r.get("dispatched_by","—"),
            fmt_dt(r.get("dispatched_at")),
            r.get("dispatch_note",""),
            r.get("replaced_by","—"),
            fmt_dt(r.get("replaced_at")),
            r.get("quantity",1),
            (r.get("status","") or "").upper(),
            (r.get("priority","") or "").upper(),
            r.get("request_id",""),
            "",
        ]

        cols = "ABCDEFGHIJKLMNOPQRST"
        for j, val in enumerate(vals):
            c = ws[f"{cols[j]}{row_num}"]
            c.value     = val
            c.font      = cell_font(bold=(j in [0,2,16]))
            c.fill      = fill(bg)
            c.alignment = left() if j in [1,4,6,7,9,10,12,13] else center()
            c.border    = thin_border()

        ws.row_dimensions[row_num].height = 18

    # ── Summary sheet ─────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.merge_cells("A1:D1")
    ws2["A1"] = "Toner Audit Summary"
    ws2["A1"].font      = Font(name="Arial", size=13, bold=True, color=C_HEADER_FG)
    ws2["A1"].fill      = fill(C_HEADER_BG)
    ws2["A1"].alignment = center()
    ws2.row_dimensions[1].height = 28

    from collections import Counter
    statuses = Counter(r.get("status","unknown") for r in rows)
    branches = Counter(r.get("branch_name","") for r in rows)

    summary_rows = [
        ("",""),
        ("OVERALL", ""),
        ("Total Requests",      len(rows)),
        ("Pending",             statuses.get("pending",0)),
        ("Approved",            statuses.get("approved",0)),
        ("Dispatched",          statuses.get("dispatched",0)),
        ("Rejected",            statuses.get("rejected",0)),
        ("",""),
        ("TOP BRANCHES BY REQUESTS",""),
    ]
    for branch, count in branches.most_common(10):
        summary_rows.append((branch, count))

    for i, (label, val) in enumerate(summary_rows, start=2):
        ws2[f"A{i}"] = label
        ws2[f"B{i}"] = val
        if label in ("OVERALL","TOP BRANCHES BY REQUESTS"):
            ws2[f"A{i}"].font = Font(name="Arial", size=10, bold=True, color=C_HEADER_FG)
            ws2[f"A{i}"].fill = fill("334155")
        ws2[f"A{i}"].font = Font(name="Arial", size=10, bold=(label in ("OVERALL","TOP BRANCHES BY REQUESTS","Total Requests")))
        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 15

    # ── Column widths (audit sheet) ───────────────────────────
    col_widths = [10, 20, 14, 18, 18, 20, 22, 18, 20, 22, 18, 20, 22, 18, 20, 6, 12, 10, 8, 4]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A6"  # freeze title + header rows

    # ── Stream output ─────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"SoftWave_Toner_Audit_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )